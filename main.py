import os
import openpyxl
import re
import sys

from itertools import islice
from tkinter import messagebox, Tk

from PyQt5 import QtWidgets
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPalette, QColor, QIcon
from PyQt5.QtWidgets import *


filepath = os.path.join(os.path.dirname(__file__))
icon_path = os.path.join(filepath, "support_files", "NukeXApp.ico")
ui_files = os.path.join(filepath, "ui_files")

sys.path.insert(0, ui_files)

from qt_nuke_template_v004 import Ui_MainWindow

__author__ = "Anandhaiyappan"
__email__ = "anandhaiyappan@gmail.com"


class NukeTemplateGen(Ui_MainWindow, QMainWindow):
    def __init__(self):
        super(NukeTemplateGen, self).__init__()
        self.setupUi(self)
        self.pb_loadnk.clicked.connect(self.load_nk)
        self.pb_load_excel = self.findChild(QtWidgets.QPushButton, "pb_loadexcel")
        self.pb_load_excel.clicked.connect(self.load_excel)
        self.pb_oppath.clicked.connect(self.load_op_path)
        self.pb_create.clicked.connect(self.script_creation)
        self.pb_clear.clicked.connect(self.ui_clear)
        self.pb_close.clicked.connect(self.close)
        self.rb_allshots.setChecked(True)
        self.le_numberofshots.setEnabled(False)
        self.rb_allshots.clicked.connect(self.enable_le)
        self.rb_selshots.clicked.connect(self.disable_le)
        self.pb_openpath.clicked.connect(self.open_path)
        self.pb_createsource.clicked.connect(self.create_source)
        self.pb_createexcel.clicked.connect(self.create_excel)
        self.setWindowIcon(QIcon(icon_path))
        self.show()
        self.op_path = ""
        self.xl_path = ""
        self.op_path_name = ""
        self.nukeName = ""
        self.excelName = ""
        self.start_frame = ""
        self.last_frame = ""
        self.shot_name = ""
        self.shot_list = dict()
        self.template_script_path = ""
        self.list_of_line_edits = list()
        self.no_of_shots = list()
        self.sheet = None

    def create_excel(self):
        try:
            name = QFileDialog.getSaveFileName(self, 'Save File', filter="*.xlsx")
            cr_xl_file = (name[0])
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sample_Sheet"
            c1 = ws.cell(row=1, column=1)
            c1.value = "Shot Name"
            c2 = ws.cell(row=1, column=2)
            c2.value = "First"
            c3 = ws.cell(row=1, column=3)
            c3.value = "Last"
            wb.save(cr_xl_file)
            t_root = Tk()
            t_root.withdraw()
            ask = messagebox.askyesno("Open Excel", "Do you want to Open Excel File?")
            if ask:
                os.startfile(cr_xl_file)
        except Exception as err:
            self.exception_handling("Error", "{}!".format(str(err)))

    def create_source(self):
        try:
            name = QFileDialog.getSaveFileName(self, 'Save File', filter="*.nk")
            nk_file = (name[0])
            if nk_file:
                with open(nk_file, 'w') as temp_nk:
                    text = ("""#! C:/Program Files/Nuke12.1v1/nuke-12.1.1.dll -nx version 12.1 v1 
                    define_window_layout_xml {<?xml version="1.0" encoding="UTF-8"?> <layout version="1.0">
                     <window	x="-1" y="-8" w="1920" h="1009" maximized="1" screen="0"> <splitter orientation="1"> 
                     <split size="40"/> <dock id="" hideTitles="1" activePageId="Toolbar.1">
                     <page id="Toolbar.1"/> </dock>
                     <split size="1257" stretch="1"/> <splitter orientation="2"> <split size="671"/> 
                     <dock id=""activePageId="uk.co.thefoundry.scripteditor.1"> <page id="Viewer.1"/> 
                     <page	id="uk.co.thefoundry.scripteditor.1"/> </dock> <split size="300"/> 
                     <dock id=""activePageId="DAG.1" focus="true"> <page id="DAG.1"/> <page id="Curve Editor.1"/> 
                     <page	id="DopeSheet.1"/> </dock> </splitter> <split size="615"/> 
                     <dock id=""activePageId="Properties.1"> <page id="Properties.1"/> 
                     <page id="uk.co.thefoundry.backgroundrenderview.1"/> </dock> </splitter> </window> </layout> }
                      Root { inputs 0 frame 1016 first_frame 1001 last_frame 1089 lock_range true fps 23.976 
                      format "4448 3096 0 0 4448 3096 1 " proxy_type scale proxy_format 
                      "1024 778 0 0 1024 778 1 1K_Super_35(full-ap)"
                      colorManagement Nuke workingSpaceLUT linear monitorLut sRGB int8Lut sRGB int16Lut sRGB logLut
                      Cineon floatLut linear onScriptSave 
                      "import re\\nnew_ver = os.path.basename(nuke.scriptName()).split(\\".nk\\")
                      \\[0]\\[-4:]\\nrt = nuke.root()\\nrd_nodes = rt\\['read_nodes'].getValue()
                      \\nwr_nodes = rt\\['write_nodes'].getValue()\\nread_list = \\[]\\nwrite_list = \\[]
                      \\nif rd_nodes:\\n    read_list = rd_nodes.strip().split(\\",\\")\\nif wr_nodes:\\n
                      write_list = wr_nodes.strip().split(\\",\\")\\nfor i in nuke.allNodes():
                      \\n    if i.Class() == \\"Write\\":  
                      \\n        if i.name() in write_list:\\n            old_value = i\\['file'].getValue()\\n
                      new_value = re.sub('\\[v]+\\[0-9]+',new_ver,old_value)\\n            i\\['file'].setValue(
                      new_value)\\n    if i.Class() == \\"Read\\":\\n        if i.name() in read_list:\\n
                      old_value = i\\['file'].getValue()\\n            new_value = re.sub('\\[v]+\\[0-9]+',new_ver,
                      old_value)\\n            i\\['file'].setValue(new_value)" addUserKnob {20 user l User}
                      addUserKnob {26 "" +STARTLINE} addUserKnob {26 _1 l "" +STARTLINE T "Type your Read and Write
                      node names (separated by comma if multiple) which need to be "} addUserKnob {26 _2 l ""
                      +STARTLINE T "autoamatically updated based on the scriptname."} addUserKnob {26 "" +STARTLINE}
                      addUserKnob {1 read_nodes l "Read Node Names" t "Put the read node names separated by comma if
                      multiple"} addUserKnob {26 "" +STARTLINE} addUserKnob {1 write_nodes l "Write Node Names" t "Put
                      the write node names separated by comma if multiple"} }	""")
                    temp_nk.write(text)
                    temp_nk.close()
                    t_root = Tk()
                    t_root.withdraw()
                    ask = messagebox.askyesno("Open Nuke", "Do you want to Open Nuke?")
                    if ask:
                        os.startfile(nk_file)
        except Exception as err:
            self.exception_handling("Error", "{}!".format(str(err)))

    def open_path(self):
        temp_path = self.le_oppath.text()
        try:
            if temp_path:
                os.startfile(temp_path)
            else:
                self.exception_handling("Error", "Invalid Path! ")
        except Exception as err:
            self.exception_handling("Error", "{}!".format(str(err)))

    def enable_le(self):
        self.le_numberofshots.setEnabled(False)
        try:
            self.le_numberofshots.setText(str(len(self.no_of_shots)))
        except Exception as err:
            self.exception_handling("Error", "{}!".format(str(err)))

    def disable_le(self):
        self.le_numberofshots.setEnabled(True)

    def load_nk(self):
        try:
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            self.nukeName, _ = QFileDialog.getOpenFileName(None, "Load Nuke File", "",
                                                           "Nuke Files (*.nk)", options=options, )

            if self.nukeName:
                self.le_nkpath.setText(self.nukeName)
                self.template_script_path = self.le_nkpath.text()
                with open(self.template_script_path) as f:
                    t_data = f.read()
                    script_check = r"onScriptSave"
                    script_match = re.search(script_check, t_data)
                f.close()

                with open(self.template_script_path) as f:
                    data = f.readlines()
                    if script_match is not None:
                        for lins in data:
                            first_regex = r'first_frame \d+'
                            first_match = re.search(first_regex, lins)
                            if first_match is not None:
                                self.le_ftframe.setText(lins.split('first_frame')[1].strip())
                            last_regex = r'last_frame \d+'
                            last_match = re.search(last_regex, lins)
                            if last_match is not None:
                                self.le_ltframe.setText(lins.split('last_frame')[1].strip())
                    else:
                        self.exception_handling("Warning", "Nuke File missing the script!")
                f.close()

        except Exception as err:
            self.exception_handling("Error", "{}!".format(str(err)))

    def load_excel(self):
        try:
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            self.excelName, _ = QFileDialog.getOpenFileName(None, "Load Excel", "",
                                                            "Excel Files (*.xlsx)", options=options, )
            if self.excelName:
                self.le_excelpath.setText(self.excelName)
                self.xl_path = self.le_excelpath.text()
                self.start_frame = self.le_ftframe.text()
                self.last_frame = self.le_ltframe.text()
                wb = openpyxl.load_workbook(self.xl_path)
                self.sheet = wb.active
                self.no_of_shots = []
                for row in self.sheet.iter_rows(min_row=2, values_only=True):
                    shot_name = row[0]
                    if shot_name:
                        try:
                            self.no_of_shots.append(shot_name)
                        except Exception as err:
                            self.exception_handling("Error", "Loaded Excel looks Invalid! {}".format(err))
                            self.le_excelpath.setText("")
                            break
                self.le_numberofshots.setText(str(len(self.no_of_shots)))
        except Exception as err:
            self.exception_handling("Error", "{}!".format(str(err)))

    def load_op_path(self):
        try:
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            self.op_path_name = QtWidgets.QFileDialog.getExistingDirectory()
            if self.op_path_name:
                self.le_oppath.setText(self.op_path_name)
        except Exception as err:
            self.exception_handling("Error", "{}!".format(str(err)))

    def ui_clear(self):
        self.list_of_line_edits = [self.le_excelpath, self.le_nkpath, self.le_oppath, self.le_shotname, self.le_ftframe,
                                   self.le_ltframe, self.le_numberofshots]
        for fields in self.list_of_line_edits:
            fields.clear()

    def script_creation(self):
        try:
            self.xl_path = self.le_excelpath.text()
            self.start_frame = self.le_ftframe.text()
            self.last_frame = self.le_ltframe.text()
            self.template_script_path = self.le_nkpath.text()
            self.shot_name = self.le_shotname.text()
            self.op_path = self.le_oppath.text()
            if self.xl_path and self.start_frame and self.last_frame and self.template_script_path and self.shot_name\
                    and self.op_path:
                if not os.path.isdir(self.op_path):
                    self.exception_handling("Error", "Invalid Path")
                    return
                if not (self.start_frame.isdigit() and self.last_frame.isdigit()):
                    self.exception_handling("Error", "First and Last Frame can have only numbers")
                    return
                self.shot_list = dict()
                for row in self.sheet.iter_rows(min_row=2, values_only=True):

                    shot_name = row[0]
                    if shot_name:
                        frame_range = [str(int(row[1])), str(int(row[2]))]
                        self.shot_list[shot_name] = frame_range
                first_frame = self.start_frame
                last_frame = self.last_frame
                shot_name = self.shot_name
                no_of_shots = int(self.le_numberofshots.text())
                shots = (list(islice(self.shot_list, no_of_shots)))
                if no_of_shots:
                    for shot in shots:
                        first = self.shot_list[shot][0]
                        last = self.shot_list[shot][1]
                        new_file = r"{}\{}\Nuke\{}_comp_v001.nk".format(self.op_path, shot, shot)
                        if not os.path.exists(os.path.dirname(new_file)):
                            os.makedirs(os.path.dirname(new_file))
                        with open(self.template_script_path) as f:
                            data = f.read()
                            data = data.replace(first_frame, str(first))
                            data = data.replace(last_frame, str(last))
                            new_data = data.replace(shot_name, shot)
                            with open(new_file, "w") as new:
                                new.write(new_data)
                            new.close()
                            f.close()
                    self.exception_handling("Success", "{} Nuke scripts are created Successfully".format(no_of_shots))
                    self.open_path()

                else:
                    self.exception_handling("Error", "Number of shots cannot be zero or blank")
            else:
                self.exception_handling("Error", "All the fields are mandatory! Please check and fill!")
        except Exception as err:
            self.exception_handling("Error", "{}".format(str(err)))

    @staticmethod
    def exception_handling(sts, messages):
        status = sts
        msg = messages
        t_root = Tk()
        t_root.withdraw()
        if sts == "Success":
            messagebox.showinfo("Success", "{}!".format(msg))
        else:
            messagebox.showerror("{}".format(status), "{}".format(msg))


try:
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    dark_palette = QPalette()
    dark_palette.setColor(QPalette.Window, QColor(53, 53, 53))
    dark_palette.setColor(QPalette.WindowText, Qt.white)
    dark_palette.setColor(QPalette.Base, QColor(35, 35, 35))
    dark_palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
    dark_palette.setColor(QPalette.ToolTipBase, QColor(25, 25, 25))
    dark_palette.setColor(QPalette.ToolTipText, Qt.white)
    dark_palette.setColor(QPalette.Text, Qt.white)
    dark_palette.setColor(QPalette.Button, QColor(53, 53, 53))
    dark_palette.setColor(QPalette.ButtonText, Qt.white)
    dark_palette.setColor(QPalette.BrightText, Qt.red)
    dark_palette.setColor(QPalette.Link, QColor(42, 130, 218))
    dark_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
    dark_palette.setColor(QPalette.HighlightedText, QColor(35, 35, 35))
    dark_palette.setColor(QPalette.Active, QPalette.Button, QColor(50, 50, 50))
    QApplication.setPalette(dark_palette)
    app.setPalette(dark_palette)
    window = NukeTemplateGen()
    app.exec_()
except Exception as e:
    root = Tk()
    root.withdraw()
    messagebox.showerror("Error", "{}!".format(str(e)))
